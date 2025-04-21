from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
import calendar

# ==== CONFIGURAÇÕES INICIAIS ====
meses_pt = {
    "January": "Janeiro", "February": "Fevereiro", "March": "Março",
    "April": "Abril", "May": "Maio", "June": "Junho",
    "July": "Julho", "August": "Agosto", "September": "Setembro",
    "October": "Outubro", "November": "Novembro", "December": "Dezembro"
}
month_name = meses_pt[datetime.now().strftime('%B')]
year = datetime.now().year
file_path = f"escala_{month_name}_{year}.xlsx"

# Carrega o arquivo
wb = load_workbook(file_path)
ws_perf = wb["Performance"]

# ==== ABA DE TRÁFEGO ====
# Remove a existente ou cria nova
if "Controle de Tráfego" in wb.sheetnames:
    ws_trafego = wb["Controle de Tráfego"]
    ws_trafego.delete_rows(1, ws_trafego.max_row)
else:
    ws_trafego = wb.create_sheet("Controle de Tráfego")

# ==== ESTILOS ====
header_fill = PatternFill("solid", fgColor="D9E1F2")
center_align = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

# ==== CABEÇALHOS ====
headers = ["Nome", "Meta", "Realizado", "% Realizado", "Dias Restantes", "Sugestão Diária"]
for col, header in enumerate(headers, 1):
    cell = ws_trafego.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
    ws_trafego.column_dimensions[get_column_letter(col)].width = 22

# ==== CONFIG DE DATAS ====
ano = datetime.now().year
mes = datetime.now().month
start_date = datetime(ano, mes, 1)
dias_no_mes = calendar.monthrange(ano, mes)[1]
hoje = datetime.now()

dias_uteis_restantes = sum(
    1 for i in range(dias_no_mes)
    if (start_date + timedelta(days=i)).weekday() < 6 and start_date + timedelta(days=i) >= hoje
)

# ==== POSIÇÕES DAS COLUNAS ====
col_nome_perf = 1
col_meta_perf = 3
col_data_inicio_perf = 8

# ==== LISTA DE VENDEDORES ====
row_start_perf = 5
vendedores = []
row = row_start_perf
while True:
    nome = ws_perf.cell(row=row, column=col_nome_perf).value
    if not nome:
        break
    vendedores.append((row, nome))
    row += 1

# ==== PREENCHIMENTO DE DADOS ====
for idx, (row_perf, nome) in enumerate(vendedores):
    row_trafego = idx + 2

    col_nome = get_column_letter(1)
    col_meta = get_column_letter(2)
    col_realizado = get_column_letter(3)
    col_percentual = get_column_letter(4)
    col_dias_restantes = get_column_letter(5)
    col_sugestao = get_column_letter(6)

    # Nome e Meta com fórmula (link)
    ws_trafego[f"{col_nome}{row_trafego}"] = f"=Performance!{get_column_letter(col_nome_perf)}{row_perf}"
    ws_trafego[f"{col_meta}{row_trafego}"] = f"=Performance!{get_column_letter(col_meta_perf)}{row_perf}"

    # Realizado = soma dinâmica dos dias preenchidos na aba Performance
    ultima_coluna_perf = ws_perf.max_column
    letra_inicio = get_column_letter(col_data_inicio_perf)
    letra_fim = get_column_letter(ultima_coluna_perf)
    ws_trafego[f"{col_realizado}{row_trafego}"] = (
        f"=SOMA(Performance!{letra_inicio}{row_perf}:{letra_fim}{row_perf})"
    )

    # % Realizado
    ws_trafego[f"{col_percentual}{row_trafego}"] = f"={col_realizado}{row_trafego}/{col_meta}{row_trafego}"

    # Dias Restantes (fixo)
    ws_trafego[f"{col_dias_restantes}{row_trafego}"] = dias_uteis_restantes

    # Sugestão Diária = (Meta - Realizado) / Dias Restantes
    ws_trafego[f"{col_sugestao}{row_trafego}"] = (
        f"=SE({col_dias_restantes}{row_trafego}>0;"
        f"({col_meta}{row_trafego}-{col_realizado}{row_trafego})/{col_dias_restantes}{row_trafego};0)"
    )

    # Aplicar estilos
    for col in range(1, len(headers) + 1):
        cell = ws_trafego.cell(row=row_trafego, column=col)
        cell.alignment = center_align
        cell.border = thin_border

# ==== SALVA ARQUIVO ====
wb.save(file_path)
print(f"✅ Planilha atualizada com sucesso: {file_path}")


""" 
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
import calendar

# Tradução de meses
meses_pt = {
    "January": "Janeiro", "February": "Fevereiro", "March": "Março", "April": "Abril",
    "May": "Maio", "June": "Junho", "July": "Julho", "August": "Agosto",
    "September": "Setembro", "October": "Outubro", "November": "Novembro", "December": "Dezembro"
}
month_name = meses_pt[datetime.now().strftime('%B')]
year = datetime.now().year
file_path = f'escala_{month_name}_{year}.xlsx'

# Abrir workbook e aba Performance
wb = load_workbook(file_path)
ws_perf = wb["Performance"]

# Verifica e limpa aba "Controle de Tráfego"
if "Controle de Tráfego" in wb.sheetnames:
    ws_trafego = wb["Controle de Tráfego"]
    ws_trafego.delete_rows(2, ws_trafego.max_row)
else:
    ws_trafego = wb.create_sheet(title="Controle de Tráfego")

# Estilos
header_fill = PatternFill("solid", fgColor="D9E1F2")
center_align = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Cabeçalhos
headers_trafego = ["Nome", "Meta", "Realizado", "% Realizado", "Dias Restantes", "Sugestão Diária"]
for col_idx, header in enumerate(headers_trafego, start=1):
    cell = ws_trafego.cell(row=1, column=col_idx, value=header)
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
    ws_trafego.column_dimensions[get_column_letter(col_idx)].width = 22

# Configs fixas
ano = datetime.now().year
mes = datetime.now().month
start_date = datetime(ano, mes, 1)
dias_no_mes = calendar.monthrange(ano, mes)[1]
hoje = datetime.now()

dias_uteis_restantes = sum(
    1 for i in range(dias_no_mes)
    if (start_date + timedelta(days=i)).weekday() < 6 and (start_date + timedelta(days=i)) >= hoje
)

# Posição das colunas na aba Performance
col_nome_perf = 1
col_meta_perf = 3
col_realizado_perf = 5  # <- Coluna onde está o total de vendas realizadas
row_start_perf = 5

# Coleta dos vendedores
vendedores = []
row = row_start_perf
while True:
    nome = ws_perf.cell(row=row, column=col_nome_perf).value
    if not nome:
        break
    vendedores.append((row, nome))
    row += 1

# Loop para preencher a aba Controle de Tráfego
for idx, (row_perf, nome) in enumerate(vendedores):
    row_trafego = idx + 2
    col_nome = get_column_letter(1)
    col_meta = get_column_letter(2)
    col_realizado = get_column_letter(3)
    col_percentual = get_column_letter(4)
    col_dias_restantes = get_column_letter(5)
    col_sugestao = get_column_letter(6)

    # Nome e Meta
    ws_trafego[f"{col_nome}{row_trafego}"] = f"='{ws_perf.title}'!{get_column_letter(col_nome_perf)}{row_perf}"
    ws_trafego[f"{col_meta}{row_trafego}"] = f"='{ws_perf.title}'!{get_column_letter(col_meta_perf)}{row_perf}"

    # Realizado: puxa direto da coluna "Performance"
    col_realizado_letra = get_column_letter(col_realizado_perf)
    ws_trafego[f"{col_realizado}{row_trafego}"] = f"='{ws_perf.title}'!{col_realizado_letra}{row_perf}"

    # % Realizado
    ws_trafego[f"{col_percentual}{row_trafego}"] = f"=SEERRO({col_realizado}{row_trafego}/{col_meta}{row_trafego};0)"

    # Dias Restantes (valor fixo)
    ws_trafego[f"{col_dias_restantes}{row_trafego}"] = dias_uteis_restantes

    # Sugestão Diária
    ws_trafego[f"{col_sugestao}{row_trafego}"] = (
        f"=SE({col_dias_restantes}{row_trafego}>0;"
        f"({col_meta}{row_trafego}-{col_realizado}{row_trafego})/{col_dias_restantes}{row_trafego};0)"
    )

    # Estilo por linha
    for col in range(1, len(headers_trafego) + 1):
        cell = ws_trafego.cell(row=row_trafego, column=col)
        cell.alignment = center_align
        cell.border = thin_border

# Salvar
wb.save(file_path)
print(f"Aba 'Controle de Tráfego' atualizada com sucesso em {file_path}")

"""