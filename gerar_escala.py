from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import calendar

# ==== LOCALE ==== (não aplicável no ambiente atual, mas funciona no Windows local)
import locale
try:
    locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')
except:
    try:
        locale.setlocale(locale.LC_TIME, 'Portuguese_Brazil.1252')
    except:
        pass  # Caso esteja num ambiente que não suporta

# ==== INPUT USUÁRIO ====
mes = int(input("Digite o número do mês (1-12): "))
ano = int(input("Digite o ano (ex: 2025): "))

# ==== CONFIGS ====
horarios = {
    "Integral": "08:00 - 17:00",
    "Manhã": "08:00 - 16:20",
    "Tarde": "09:40 - 18:00"
}

Colaboradores = [
    {"Nome":"Alando", "Cargo":"Gerência", "Horário:": "Integral"},
    {"Nome":"Cassia", "Cargo":"Gerência", "Horário:": "Manhã"},
    {"Nome":"Pedro", "Cargo":"Repositor - Auxiliar", "Horário:": "Manhã"},
    {"Nome":"Vanessa", "Cargo":"Vendedor(a)", "Horário:": "Manhã"},
    {"Nome":"Junior", "Cargo":"Vendedor(a)", "Horário:": "Manhã"},
    {"Nome":"Ronald", "Cargo":"Vendedor(a)", "Horário:": "Manhã"},
    {"Nome":"Manoel", "Cargo":"Vendedor(a)", "Horário:": "Manhã"},
    {"Nome":"Michel", "Cargo":"Vendedor(a)", "Horário:": "Manhã"},
    {"Nome":"Richard", "Cargo":"Vendedor(a)", "Horário:": "Tarde"},
    {"Nome":"Alexandre", "Cargo":"Vendedor(a)", "Horário:": "Tarde"},
    {"Nome":"Aurélio", "Cargo":"Vendedor(a)", "Horário:": "Tarde"},
    {"Nome":"Bruno", "Cargo":"Vendedor(a)", "Horário:": "Tarde"},
    {"Nome":"Isabella", "Cargo":"Operador(a) SAC", "Horário:": "Manhã"},
    {"Nome":"Camila", "Cargo":"Operador(a) SAC", "Horário:": "Tarde"},
    {"Nome":"Luciene", "Cargo":"Operador(a) Caixa", "Horário:": "Manhã"},
    {"Nome":"Victor", "Cargo":"Operador(a) Caixa", "Horário:": "Tarde"},
    {"Nome":"Esther", "Cargo":"Operador(a) Caixa", "Horário:": "Tarde"},
    {"Nome":"Diego", "Cargo":"Repositor de Loja", "Horário:": "Manhã"},
    {"Nome":"Thiago", "Cargo":"Repositor de Loja", "Horário:": "Tarde"},
]

# Atualiza campo “Horário:” com faixa horária
for c in Colaboradores:
    horario_tipo = c.get("Horário:", "")
    c["Horário:"] = horarios.get(horario_tipo, "")

# Agrupa por cargo, e dentro de cada cargo ordena por turno (Integral → Manhã → Tarde)
def ordem_por_cargo_e_turno(c):
    turno = next((k for k, v in horarios.items() if v == c["Horário:"]), "")
    prioridade_turno = {"Integral": 0, "Manhã": 1, "Tarde": 2}
    return (c["Cargo"], prioridade_turno.get(turno, 99), c["Nome"])

Colaboradores.sort(key=ordem_por_cargo_e_turno)

# ==== ESTILOS ====
header_fill = PatternFill("solid", fgColor="D9E1F2")
weekend_fill = PatternFill("solid", fgColor="FCE4D6")
total_fill = PatternFill("solid", fgColor="E2EFDA")
center_align = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ==== PLANILHA ====
wb = Workbook()
ws = wb.active
ws.title = "Escala de Plantão"

headers = ["Nome", "Cargo", "Horário:"]
dias_no_mes = calendar.monthrange(ano, mes)[1]
start_date = datetime(ano, mes, 1)


# Conta a quantidade de dias úteis (segunda a sábado)
dias_uteis = sum(1 for i in range(dias_no_mes) 
                 if (start_date + timedelta(days=i)).weekday() < 6)


inicio_coluna = len(headers) + 1
row_data = 3
row_dow = 4
row_start_names = 5
row_total = row_start_names + len(Colaboradores)

# Cabeçalhos
for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col_idx)].width = 22

# Dados dos colaboradores
for row_idx, colaborador in enumerate(Colaboradores, start=row_start_names):
    for col_idx, header in enumerate(headers, start=1):
        valor = colaborador.get(header, "")
        cell = ws.cell(row=row_idx, column=col_idx, value=valor)
        cell.alignment = center_align
        cell.border = thin_border

# Datas e dias da semana
# Lista de dias úteis (segunda a sábado)
dias_uteis_plantao = []
for dia in range(dias_no_mes):
    data = start_date + timedelta(days=dia)
    if data.weekday() < 6:  # segunda (0) a sábado (5)
        dias_uteis_plantao.append(data)

# Cabeçalhos com datas e dias úteis
for idx, current_date in enumerate(dias_uteis_plantao):
    col = get_column_letter(inicio_coluna + idx)
    data_str = current_date.strftime("%d/%b")
    dias_semana_pt = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]
    dia_semana = dias_semana_pt[current_date.weekday()]
    is_weekend = current_date.weekday() >= 5

    ws[f"{col}{row_data}"] = data_str
    ws[f"{col}{row_dow}"] = dia_semana

    for r in [row_data, row_dow]:
        cell = ws[f"{col}{r}"]
        cell.fill = weekend_fill if is_weekend else header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Células de plantão
    for row in range(row_start_names, row_start_names + len(Colaboradores)):
        cell = ws[f"{col}{row}"]
        cell.value = ""
        cell.alignment = center_align
        cell.border = thin_border

    # Fórmula de total
    total_cell = ws[f"{col}{row_total}"]
    total_cell.value = f'=COUNTIF({col}{row_start_names}:{col}{row_total - 1}, "Sim")'
    total_cell.font = Font(bold=True)
    total_cell.alignment = center_align
    total_cell.fill = total_fill
    total_cell.border = thin_border


# Linha de total
ws[f"A{row_total}"] = "Total por dia"
ws[f"A{row_total}"].font = Font(bold=True)
ws[f"A{row_total}"].alignment = center_align
ws[f"A{row_total}"].fill = total_fill
ws[f"A{row_total}"].border = thin_border


# ==== ABA PERFORMANCE COM VENDEDOR(A) E DIAS ÚTEIS ====
ws_perf = wb.create_sheet(title="Performance")

headers = ["Nome","Horário:", "Meta", "Dias Trabalhados", "Performance", "Nova Necessidade", "%"]
inicio_coluna_perf = len(headers) + 1
row_data_perf = 3
row_dow_perf = 4
row_start_perf = 5

# Cabeçalhos fixos
for col_idx, header in enumerate(headers, start=1):
    cell = ws_perf.cell(row=1, column=col_idx, value=header)
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
    ws_perf.column_dimensions[get_column_letter(col_idx)].width = 22
    
# Dias úteis (segunda a sábado)
dias_uteis = []
total_dias_uteis = 0
for dia in range(dias_no_mes):
    current_date = start_date + timedelta(days=dia)
    if current_date.weekday() < 6:  # segunda (0) a sábado (5)
        dias_uteis.append((dia, current_date))
        total_dias_uteis += 1


# Cabeçalhos de datas e dias da semana
for i, (dia_idx, current_date) in enumerate(dias_uteis):
    col = get_column_letter(inicio_coluna_perf + i)
    data_str = current_date.strftime("%d/%b")
    dias_semana_pt = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]
    dia_semana = dias_semana_pt[current_date.weekday()]


    for row_idx, value in [(row_data_perf, data_str), (row_dow_perf, dia_semana)]:
        cell = ws_perf[f"{col}{row_idx}"]
        cell.value = value
        cell.fill = weekend_fill if current_date.weekday() >= 5 else header_fill
        cell.alignment = center_align
        cell.border = thin_border

# Filtrar vendedores
vendedores = [c for c in Colaboradores if c["Cargo"] == "Vendedor(a)"]

# Preenche dados dos vendedores
for row_offset, vendedor in enumerate(vendedores):
    row = row_start_perf + row_offset
    for col_idx, header in enumerate(headers, start=1):
        valor = vendedor.get(header, "")
        cell = ws_perf.cell(row=row, column=col_idx, value=valor)
        cell.alignment = center_align
        cell.border = thin_border

    # Células de escala
    for i in range(len(dias_uteis)):
        col = get_column_letter(inicio_coluna_perf + i)
        cell = ws_perf[f"{col}{row}"]
        cell.value = ""
        cell.alignment = center_align
        cell.border = thin_border

# Linha total por dia
row_total_perf = row_start_perf + len(vendedores)
ws_perf[f"A{row_total_perf + 2}"] = f"Total de dias úteis: {total_dias_uteis}"
ws_perf[f"A{row_total_perf + 2}"].font = Font(bold=True)
ws_perf[f"A{row_total_perf}"].alignment = center_align
ws_perf[f"A{row_total_perf}"].fill = total_fill
ws_perf[f"A{row_total_perf}"].border = thin_border

# Fórmulas de total por coluna
for i in range(len(dias_uteis)):
    col = get_column_letter(inicio_coluna_perf + i)
    total_cell = ws_perf[f"{col}{row_total_perf}"]
    total_cell.value = f'=COUNTIF({col}{row_start_perf}:{col}{row_total_perf - 1}, "Sim")'
    total_cell.font = Font(bold=True)
    total_cell.alignment = center_align
    total_cell.fill = total_fill
    total_cell.border = thin_border
    
    

# Salva
nome_mes = start_date.strftime('%B').capitalize()
file_path = f"./escala_{nome_mes}_{ano}.xlsx"
wb.save(file_path)
print(f"Planilha salva em: {file_path}")
